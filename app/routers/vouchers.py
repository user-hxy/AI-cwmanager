"""路由 - 凭证管理（录入、审核、过账）"""
from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import RedirectResponse, JSONResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime
from app.database import get_db
from app.models import Voucher, VoucherEntry, Account, User
from app.routers.auth import get_login_user, get_user_from_cookie, templates, require_active_company
from app.services.voucher_service import generate_voucher_no, post_voucher
from app.routers.settings import get_setting
from app.services.period_helper import get_target_period, get_current_period
from app.models.misc import AuditLog, SceneRule, BankReceipt, Invoice

router = APIRouter(prefix="/vouchers", tags=["凭证管理"])


@router.get("/")
async def voucher_list(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    if not user.company_id:
        return RedirectResponse(url="/dashboard", status_code=302)

    # 优先定位到有待处理凭证的月份（草稿/待审），否则用 get_current_period
    default_period = get_current_period(request, db, user.company_id)
    latest_unprocessed = db.query(
        func.strftime("%Y-%m", Voucher.date).label("p")
    ).filter(
        Voucher.company_id == user.company_id,
        Voucher.status.in_(["draft", "pending"]),
    ).order_by(func.strftime("%Y-%m", Voucher.date).desc()).first()
    if latest_unprocessed and request.query_params.get("period"):
        # 如果URL有参数，用URL的；否则自动定位
        pass
    if latest_unprocessed and "period" not in request.query_params:
        default_period = latest_unprocessed[0]

    period = request.query_params.get("period", default_period)
    period_start = request.query_params.get("period_start", period)
    period_end = request.query_params.get("period_end", period)
    status_filter = request.query_params.get("status", "")
    query = db.query(Voucher).filter(Voucher.company_id == user.company_id)
    # 期间范围过滤
    if period_start and period_end:
        query = query.filter(
            func.strftime("%Y-%m", Voucher.date) >= period_start,
            func.strftime("%Y-%m", Voucher.date) <= period_end,
        )
    else:
        query = query.filter(func.strftime("%Y-%m", Voucher.date) == period)
    if status_filter:
        query = query.filter(Voucher.status == status_filter)
    vouchers = query.order_by(Voucher.voucher_no).all()

    # 获取每个凭证的借方金额合计（作为凭证金额展示）
    debit_totals = {}
    if vouchers:
        for v in vouchers:
            total = db.query(func.sum(VoucherEntry.amount)).filter(
                VoucherEntry.voucher_id == v.id,
                VoucherEntry.direction == "借",
            ).scalar()
            debit_totals[v.id] = float(total) if total else 0.0
    
    # 获取各期间的草稿数量（供前端提示）
    draft_periods = db.query(
        func.strftime("%Y-%m", Voucher.date).label("p"),
        func.count(Voucher.id).label("c")
    ).filter(
        Voucher.company_id == user.company_id,
        Voucher.status == "draft",
    ).group_by(func.strftime("%Y-%m", Voucher.date)).order_by(
        func.strftime("%Y-%m", Voucher.date).desc()
    ).all()
    accounts = db.query(Account).filter(Account.company_id == user.company_id).order_by(Account.code).all()
    return templates(request, "vouchers/list.html", {
        "vouchers": vouchers, "user": user, "period": period,
        "period_start": period_start, "period_end": period_end,
        "status_filter": status_filter, "accounts": accounts,
        "draft_periods": draft_periods, "debit_totals": debit_totals,
    })


@router.get("/suggest-accounts")
async def suggest_accounts(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """根据摘要关键词推荐科目"""
    if not user.company_id:
        return JSONResponse({"accounts": []})
    summary = request.query_params.get("summary", "")
    if not summary:
        return JSONResponse({"accounts": []})

    amount = request.query_params.get("amount", "0")
    tax_rate = request.query_params.get("tax_rate", "0")
    total_amt = float(amount) if amount else 0
    tax_rate_val = float(tax_rate) if tax_rate else 0
    is_invoice_scene = any(kw in summary for kw in ["开发票", "开票", "发票", "销项", "进项", "专票", "采购", "购货"])

    # 1. 场景规则（从SceneRule数据库表读取，包含内置+自定义）
    from app.services.scene_service import get_scene_rules
    scene_rules = get_scene_rules(user.company_id, db)

    matched_scene = None
    for rule in scene_rules:
        if not rule.keywords:
            continue
        kw_list = [k.strip() for k in rule.keywords.split(",")]
        if any(k in summary for k in kw_list if k):
            debit_acct = db.query(Account).filter(
                Account.company_id == user.company_id, Account.code.like(f"{rule.debit_account_code}%"),
            ).first()
            credit_acct = db.query(Account).filter(
                Account.company_id == user.company_id, Account.code.like(f"{rule.credit_account_code}%"),
            ).first()
            if debit_acct and credit_acct:
                matched_scene = (debit_acct, credit_acct)
                break

    if matched_scene:
        debit_acct, credit_acct = matched_scene
        unique = [
            {"account_id": debit_acct.id, "direction": "借"},
            {"account_id": credit_acct.id, "direction": "贷"},
        ]
    else:
        # 无匹配场景时，按原有关键词映射规则匹配
        suggestions = []
        builtin_rules = [
            (["办公用品", "办公费", "文具", "耗材", "打印"], "5602", "借"),
            (["差旅费", "出差", "交通", "住宿", "机票", "火车"], "5602", "借"),
            (["计提工资", "计提薪酬", "计提社保", "计提公积金", "计提奖金"], "2211", "贷"),
            (["发放工资", "发工资", "支付工资", "工资发放"], "2211", "借"),
            (["工资", "薪酬"], "5602", "借"),
            (["税", "缴税", "增值税", "所得税", "印花税"], "2221", "借"),
            (["采购", "购货", "商品", "库存", "材料", "原材料"], "1405", "借"),
            (["保险", "车辆险", "财产险"], "5602", "借"),
            (["其他费用", "杂费", "服务费", "咨询费"], "5602", "借"),
            (["租金", "房租", "物业"], "5602", "借"),
            (["银行", "手续费", "利息", "汇款"], "5603", "借"),
            (["销售", "销售收入", "营业收入", "货款"], "5001", "贷"),
            (["收款", "收回应收"], "1122", "贷"),
            (["投资", "分红", "股利"], "5111", "贷"),
            (["固定资产", "设备", "机器"], "1601", "借"),
            (["客户", "应收", "欠款", "未收款"], "1122", "借"),
            (["供应商", "应付", "欠款"], "2202", "贷"),
        ]
        for keywords, code, direction in builtin_rules:
            if any(k in summary for k in keywords):
                if not any(s.get("account_code") == code for s in suggestions):
                    acct = db.query(Account).filter(
                        Account.company_id == user.company_id, Account.code.like(f"{code}%"),
                    ).first()
                    if acct:
                        suggestions.append({"account_id": acct.id, "direction": direction})

        seen = set()
        unique = []
        for s in suggestions:
            key = f"{s['account_id']}_{s['direction']}"
            if key not in seen:
                seen.add(key)
                unique.append(s)

        if len(unique) == 1:
            first = unique[0]
            cash_acct = db.query(Account).filter(
                Account.company_id == user.company_id, Account.code.like("1002%"),
            ).first()
            if cash_acct:
                opposite_dir = "贷" if first['direction'] == "借" else "借"
                if f"{cash_acct.id}_{opposite_dir}" not in seen:
                    unique.append({"account_id": cash_acct.id, "direction": opposite_dir})

    # 推断凭证字
    voucher_word = "转"
    # 计提类 → 转
    if any(kw in summary for kw in ["计提", "摊销", "折旧", "结转"]):
        voucher_word = "转"
    elif any(kw in summary for kw in [
        "收入", "收款", "货款", "销售", "投资", "分红",
        "收回应收", "收到", "回款", "利息收入", "银行利息",
        "存款利息", "利息入账",
        "实收资本", "注资", "增资", "补贴", "补助",
        "归还", "还贷", "收回",
    ]):
        voucher_word = "收"
    elif any(kw in summary for kw in [
        "付款", "缴税", "购买", "采购", "费用", "差旅", "工资",
        "社保", "办公", "租金", "手续费", "利息", "保险",
        "报销", "接待", "借款", "罚款", "公积金",
        "扣缴", "发放", "培训", "福利", "广告", "推广",
        "快递", "维修", "加油", "电话", "网络", "通信",
        "保证金", "押金", "外包", "托管",
    ]):
        voucher_word = "付"

    # 金额处理：如果有传入金额，分配到各分录行
    if total_amt > 0 and unique:
        if is_invoice_scene and tax_rate_val > 0:
            # 发票场景：拆分不含税金额和税额
            tax_amt = round(total_amt * tax_rate_val / (100 + tax_rate_val), 2)
            net_amt = round(total_amt - tax_amt, 2)
            # 查找应交税费科目
            tax_acct = db.query(Account).filter(
                Account.company_id == user.company_id, Account.code.like("2221%"),
            ).first()
            if tax_acct and len(unique) >= 2:
                # 判断是销项(开发票)还是进项(采购)：
                # - 销项：第一条分录为借:应收账款(1122) 或 贷:收入(5xxx/6xxx)
                # - 进项：第一条分录为借:费用/成本/库存(1xxx/4xxx/5xxx)
                first_acct_code = db.query(Account).filter(Account.id == unique[0]["account_id"]).first()
                first_code = first_acct_code.code if first_acct_code else ""
                # 销项特征：第一条是应收账款(1122)借方 或 第一条是收入/营业外(5xxx/6xxx)贷方
                is_sales = (first_code.startswith("1122") and unique[0]["direction"] == "借") or first_code.startswith(("5001", "5051", "5111", "5301"))
                if is_sales:
                    # 销项（开发票）：借：应收账款(total) / 贷：收入(net) + 贷：税额(tax)
                    unique = [
                        {"account_id": unique[0]["account_id"], "direction": "借", "amount": total_amt},
                        {"account_id": unique[1]["account_id"], "direction": "贷", "amount": net_amt},
                        {"account_id": tax_acct.id, "direction": "贷", "amount": tax_amt},
                    ]
                else:
                    # 进项（采购）：借：成本/费用(net) + 借：税额(tax) / 贷：银行/应付(total)
                    entries = [
                        {"account_id": unique[0]["account_id"], "direction": "借", "amount": net_amt},
                        {"account_id": tax_acct.id, "direction": "借", "amount": tax_amt},
                    ]
                    # 贷方：保留原匹配的第二个科目（银行或应付）
                    second_id = unique[1]["account_id"] if len(unique) > 1 else None
                    if not second_id:
                        cash = db.query(Account).filter(
                            Account.company_id == user.company_id, Account.code.like("1002%"),
                        ).first()
                        second_id = cash.id if cash else 0
                    entries.append({"account_id": second_id, "direction": "贷", "amount": total_amt})
                    unique = entries
            else:
                for item in unique:
                    item["amount"] = total_amt
        else:
            # 非发票场景：金额填入所有分录
            for item in unique:
                item["amount"] = total_amt

    return JSONResponse({"accounts": unique[:5], "voucher_word": voucher_word})





@router.get("/scenes")
async def get_scenes(request: Request, db: Session = Depends(get_db)):
    """获取场景配置（前端动态加载），常用场景按企业使用频率动态排序"""
    from app.services.scene_service import get_scene_groups, get_frequent_scenes

    user = get_user_from_cookie(request, db)

    groups = {}
    frequent = []
    if user and user.company_id:
        groups = get_scene_groups(user.company_id, db)
        frequent = get_frequent_scenes(user.company_id, db)

    # 读取企业默认税率
    default_tax = get_setting(db, user.company_id, "default_tax_rate", "6") if user and user.company_id else "6"

    return JSONResponse({
        "groups": groups,
        "frequent": frequent,
        "default_tax_rate": default_tax,
    })


@router.get("/add")
async def voucher_add_page(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return RedirectResponse(url="/vouchers/", status_code=302)
    accounts = db.query(Account).filter(Account.company_id == user.company_id).order_by(Account.code).all()
    from datetime import date
    target_period = get_target_period(db, user.company_id)
    y, m = int(target_period[:4]), int(target_period[5:7])
    default_date = date(y, m, 1).isoformat()
    return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "edit_mode": False, "today": default_date, "default_entries": int(get_setting(db, user.company_id, "default_entries", "5"))})


@router.post("/add")
async def voucher_add(
    request: Request,
    date_str: str = Form(...),
    summary: str = Form(""),
    voucher_word: str = Form("记"),
    db: Session = Depends(get_db),
    user: User = Depends(require_active_company),
):
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return RedirectResponse(url="/vouchers/", status_code=302)

    form = await request.form()
    voucher_date = date.fromisoformat(date_str)

    # 解析分录
    entries_data = []
    i = 0
    while f"account_id_{i}" in form:
        account_id = int(form.get(f"account_id_{i}"))
        direction = form.get(f"direction_{i}")
        amount = float(form.get(f"amount_{i}", 0))
        entry_summary = form.get(f"entry_summary_{i}", "")
        if account_id and amount > 0:
            entries_data.append({
                "account_id": account_id, "direction": direction,
                "amount": amount, "summary": entry_summary,
            })
        i += 1

    if not entries_data:
        accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
        return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "edit_mode": False, "error": "至少需要一条分录", "today": date_str, "default_entries": int(get_setting(db, user.company_id, "default_entries", "5"))})

    # 检查借贷平衡
    debit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "借")
    credit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "贷")
    if abs(debit_sum - credit_sum) > 0.01:
        accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
        return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "edit_mode": False, "error": f"借贷不平衡: 借={debit_sum:.2f}, 贷={credit_sum:.2f}", "today": date_str, "default_entries": int(get_setting(db, user.company_id, "default_entries", "5"))})

    # 检查是否已存在计提工资/社保凭证（相同摘要关键字）
    period = voucher_date.strftime("%Y-%m")
    if "计提" in summary:
        existing = db.query(Voucher).filter(
            Voucher.company_id == user.company_id,
            Voucher.summary.contains(summary[:6]),
            func.strftime("%Y-%m", Voucher.date) == period,
        ).first()
        if existing:
            accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
            return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "edit_mode": False, "error": f"本期已存在「{summary}」凭证，请勿重复计提", "today": date_str, "default_entries": int(get_setting(db, user.company_id, "default_entries", "5"))})

    # 检查来源（发票/回单）
    invoice_id = form.get("invoice_id", "")
    receipt_id = form.get("receipt_id", "")
    if invoice_id:
        source_type = "invoice"
        source_id = int(invoice_id)
    elif receipt_id:
        source_type = "bank_receipt"
        source_id = int(receipt_id)
    else:
        source_type = "manual"
        source_id = None

    voucher_no, seq = generate_voucher_no(db, user.company_id, voucher_word, voucher_date.year, voucher_date.month)
    att_count = int(form.get("attachment_count", 0))
    voucher = Voucher(
        company_id=user.company_id, voucher_no=voucher_no,
        date=voucher_date, voucher_word=voucher_word, serial_no=seq,
        summary=summary, status="draft", source_type=source_type,
        source_id=source_id, creator_id=user.id,
        attachment_count=att_count,
    )
    db.add(voucher)
    db.flush()

    # 标记来源单据为已处理
    if source_id:
        if source_type == "invoice":
            doc = db.query(Invoice).filter(Invoice.id == source_id).first()
            if doc:
                doc.status = "processed"
                doc.voucher_id = voucher.id
        elif source_type == "bank_receipt":
            doc = db.query(BankReceipt).filter(BankReceipt.id == source_id).first()
            if doc:
                doc.status = "processed"
                doc.voucher_id = voucher.id

    for idx, entry in enumerate(entries_data):
        acct = db.query(Account).filter(Account.id == entry["account_id"]).first()
        db.add(VoucherEntry(
            voucher_id=voucher.id,
            account_id=entry["account_id"],
            account_code=acct.code if acct else "",
            account_name=acct.name if acct else "",
            direction=entry["direction"],
            amount=entry["amount"],
            summary=entry["summary"],
            sort_order=idx,
        ))
    db.commit()
    period = voucher_date.strftime("%Y-%m")
    return RedirectResponse(url=f"/vouchers/?period={period}", status_code=302)


@router.get("/{voucher_id}/edit")
async def voucher_edit_page(voucher_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """编辑凭证页面（未审核的凭证所有人都可编辑）"""
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id, Voucher.company_id == user.company_id).first()
    if not voucher:
        return RedirectResponse(url="/vouchers/", status_code=302)
    if voucher.status in ("approved", "posted"):
        return RedirectResponse(url="/vouchers/", status_code=302)

    accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
    entries = db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).order_by(VoucherEntry.sort_order).all()
    from datetime import date
    return templates(request, "vouchers/form.html", {
        "user": user, "accounts": accounts,
        "voucher": voucher, "entries": entries,
        "edit_mode": True, "today": date.today().isoformat(),
    })


@router.post("/{voucher_id}/edit")
async def voucher_edit(
    voucher_id: int,
    request: Request,
    date_str: str = Form(...),
    summary: str = Form(""),
    voucher_word: str = Form("记"),
    db: Session = Depends(get_db),
    user: User = Depends(require_active_company),
):
    """保存编辑的凭证"""
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id, Voucher.company_id == user.company_id).first()
    if not voucher:
        return RedirectResponse(url="/vouchers/", status_code=302)
    if voucher.status in ("approved", "posted"):
        return RedirectResponse(url="/vouchers/", status_code=302)

    form = await request.form()
    voucher_date = date.fromisoformat(date_str)

    entries_data = []
    i = 0
    while f"account_id_{i}" in form:
        account_id = int(form.get(f"account_id_{i}"))
        direction = form.get(f"direction_{i}")
        amount = float(form.get(f"amount_{i}", 0))
        entry_summary = form.get(f"entry_summary_{i}", "")
        if account_id and amount > 0:
            entries_data.append({
                "account_id": account_id, "direction": direction,
                "amount": amount, "summary": entry_summary,
            })
        i += 1

    if not entries_data:
        accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
        entries = db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).order_by(VoucherEntry.sort_order).all()
        return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "voucher": voucher, "entries": entries, "edit_mode": True, "error": "至少需要一条分录"})

    debit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "借")
    credit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "贷")
    if abs(debit_sum - credit_sum) > 0.01:
        accounts = db.query(Account).filter(Account.company_id == user.company_id, Account.is_detail == True).order_by(Account.code).all()
        entries = db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).order_by(VoucherEntry.sort_order).all()
        return templates(request, "vouchers/form.html", {"user": user, "accounts": accounts, "voucher": voucher, "entries": entries, "edit_mode": True, "error": f"借贷不平衡: 借={debit_sum:.2f}, 贷={credit_sum:.2f}"})

    # 如果凭证字变了，重新生成凭证号
    if voucher_word != voucher.voucher_word:
        new_no, new_seq = generate_voucher_no(db, user.company_id, voucher_word, voucher_date.year, voucher_date.month)
        voucher.voucher_word = voucher_word
        voucher.voucher_no = new_no
        voucher.serial_no = new_seq

    voucher.date = voucher_date
    voucher.summary = summary
    voucher.attachment_count = int(form.get("attachment_count", 0))
    if voucher.status == "pending":
        voucher.status = "draft"  # 编辑已提交的凭证会退回草稿状态
    db.flush()

    # 删除旧分录，重新创建
    db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).delete()
    for idx, entry in enumerate(entries_data):
        acct = db.query(Account).filter(Account.id == entry["account_id"]).first()
        db.add(VoucherEntry(
            voucher_id=voucher.id,
            account_id=entry["account_id"],
            account_code=acct.code if acct else "",
            account_name=acct.name if acct else "",
            direction=entry["direction"],
            amount=entry["amount"],
            summary=entry["summary"],
            sort_order=idx,
        ))
    db.commit()
    period = voucher_date.strftime("%Y-%m")
    return RedirectResponse(url=f"/vouchers/?period={period}", status_code=302)


@router.post("/{voucher_id}/submit")
async def submit_voucher(voucher_id: int, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """提交审核"""
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id, Voucher.company_id == user.company_id).first()
    if voucher and voucher.status == "draft" and voucher.creator_id == user.id:
        # 检查借贷平衡
        entries = db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).all()
        db_sum = sum(e.amount for e in entries if e.direction == "借")
        cr_sum = sum(e.amount for e in entries if e.direction == "贷")
        if abs(db_sum - cr_sum) > 0.01:
            return JSONResponse({"success": False, "msg": "借贷不平衡"})
        voucher.status = "pending"
        db.commit()
    period = voucher.date.strftime("%Y-%m") if 'voucher' in dir() else ""
    return RedirectResponse(url=f"/vouchers/?period={period}", status_code=302)


@router.get("/review")
async def review_list(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """审核列表"""
    if user.role not in ("reviewer", "company_admin"):
        return RedirectResponse(url="/vouchers/", status_code=302)
    vouchers = db.query(Voucher).filter(
        Voucher.company_id == user.company_id,
        Voucher.status == "pending",
    ).order_by(Voucher.date).all()
    msg = request.query_params.get("msg", "")
    return templates(request, "vouchers/review.html", {"vouchers": vouchers, "user": user, "msg": msg})


@router.post("/{voucher_id}/approve")
async def approve_voucher(voucher_id: int, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """审核通过"""
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id, Voucher.company_id == user.company_id).first()
    if voucher and voucher.status == "pending" and user.role in ("reviewer", "company_admin"):
        if voucher.creator_id == user.id:
            return JSONResponse({"success": False, "msg": "制单人与审核人不能为同一人"})
        voucher.status = "approved"
        voucher.reviewer_id = user.id
        voucher.reviewed_at = datetime.now()
        db.commit()
    return RedirectResponse(url="/vouchers/review", status_code=302)


@router.post("/batch-approve")
async def batch_approve(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """批量审核通过"""
    if user.role not in ("reviewer", "company_admin"):
        return RedirectResponse(url="/vouchers/review", status_code=302)
    form = await request.form()
    ids = form.getlist("voucher_ids")
    success = 0
    for vid in ids:
        voucher = db.query(Voucher).filter(Voucher.id == int(vid), Voucher.status == "pending").first()
        if voucher and voucher.creator_id != user.id:
            voucher.status = "approved"
            voucher.reviewer_id = user.id
            voucher.reviewed_at = datetime.now()
            success += 1
    db.commit()
    return RedirectResponse(url=f"/vouchers/review?msg=批量通过{success}张凭证", status_code=302)


@router.post("/{voucher_id}/reject")
async def reject_voucher(voucher_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """驳回"""
    form = await request.form()
    reason = form.get("reason", "")
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if voucher and voucher.status == "pending" and user.role in ("reviewer", "company_admin"):
        voucher.status = "draft"
        voucher.reject_reason = reason
        voucher.reviewer_id = None
        db.commit()
    return RedirectResponse(url="/vouchers/review", status_code=302)


@router.post("/{voucher_id}/post")
async def post_voucher_route(voucher_id: int, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """过账"""
    ok, msg = post_voucher(voucher_id, user.id, db)
    return JSONResponse({"success": ok, "msg": msg})


@router.get("/batch-post")
async def batch_post_page(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """批量过账页面"""
    if not user.company_id or user.role not in ("inputer", "company_admin", "super_admin"):
        return RedirectResponse(url="/vouchers/", status_code=302)
    period = request.query_params.get("period", "")
    if not period:
        from app.services.period_helper import get_target_period
        period = get_target_period(db, user.company_id)
    # 查出所有待过账（已审核）的凭证
    from datetime import date
    vouchers = db.query(Voucher).filter(
        Voucher.company_id == user.company_id,
        Voucher.status == "approved",
        func.strftime("%Y-%m", Voucher.date) == period,
    ).order_by(Voucher.date, Voucher.voucher_no).all()
    return templates(request, "vouchers/batch_post.html", {
        "user": user, "vouchers": vouchers, "period": period,
    })


@router.post("/batch-post")
async def batch_post(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """批量过账"""
    form = await request.form()
    ids = form.get("ids", "")
    success_count = 0
    for vid in ids.split(","):
        if vid.strip():
            ok, _ = post_voucher(int(vid.strip()), user.id, db)
            if ok:
                success_count += 1
    return JSONResponse({"success": True, "count": success_count})


@router.post("/{voucher_id}/delete")
async def delete_voucher(voucher_id: int, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """删除凭证"""
    if user.role not in ("company_admin", "super_admin", "reviewer", "inputer"):
        return JSONResponse({"success": False, "msg": "无删除权限"})
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id, Voucher.company_id == user.company_id).first()
    if not voucher:
        return JSONResponse({"success": False, "msg": "凭证不存在"})
    if voucher.status not in ("draft", "pending"):
        return JSONResponse({"success": False, "msg": "只能删除草稿或已提交状态的凭证"})

    # 先解除关联引用（bank_receipts、invoices 等表中的 voucher_id）
    db.query(BankReceipt).filter(BankReceipt.voucher_id == voucher.id).update(
        {"voucher_id": None}
    )
    db.query(Invoice).filter(Invoice.voucher_id == voucher.id).update(
        {"voucher_id": None}
    )
    # 删除凭证分录
    db.query(VoucherEntry).filter(VoucherEntry.voucher_id == voucher.id).delete()
    db.delete(voucher)
    db.add(AuditLog(
        company_id=user.company_id, user_id=user.id,
        username=user.username, action="delete_voucher",
        target_type="voucher", target_id=voucher_id,
        detail=f"删除凭证: {voucher.voucher_no}",
    ))
    db.commit()
    return JSONResponse({"success": True, "msg": "删除成功"})


@router.get("/export-pdf")
async def export_voucher_pdf(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """导出当期凭证PDF"""
    if not user.company_id:
        return RedirectResponse(url="/vouchers/", status_code=302)
    period = request.query_params.get("period", "")
    if not period:
        period = get_target_period(db, user.company_id)
    from app.services.voucher_pdf import generate_voucher_pdf
    from urllib.parse import quote
    orientation = get_setting(db, user.company_id, "voucher_orientation", "L")
    per_page = int(get_setting(db, user.company_id, "voucher_per_page", "2"))
    buf = generate_voucher_pdf(user.company_id, period, db, orientation, per_page)
    filename = f"vouchers_{period}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"{filename}\"; filename*=UTF-8''{quote(f'凭证_{period}.pdf')}"},
    )


@router.get("/{voucher_id}/detail")
async def voucher_detail(voucher_id: int, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """获取凭证详情（含分录、附件文件信息）"""
    voucher = db.query(Voucher).filter(
        Voucher.id == voucher_id, Voucher.company_id == user.company_id
    ).first()
    if not voucher:
        return JSONResponse({"success": False, "msg": "凭证不存在"})

    entries = db.query(VoucherEntry).filter(
        VoucherEntry.voucher_id == voucher.id
    ).order_by(VoucherEntry.sort_order).all()

    creator = db.query(User).filter(User.id == voucher.creator_id).first()
    reviewer = db.query(User).filter(User.id == voucher.reviewer_id).first()

    # 查找关联的原始文件（银行回单/发票）
    from app.models.misc import Attachment, BankReceipt, Invoice
    source_file = None
    if voucher.source_type == "bank_receipt" and voucher.source_id:
        att = db.query(Attachment).filter(
            Attachment.source_type == "bank_receipt",
        ).order_by(Attachment.upload_time.desc()).first()
        if att:
            from pathlib import Path
            if Path(att.file_path).exists():
                source_file = {
                    "url": f"/bank-receipts/{voucher.source_id}/file",
                    "name": att.file_name,
                }
    elif voucher.source_type == "invoice" and voucher.source_id:
        inv = db.query(Invoice).filter(Invoice.id == voucher.source_id).first()
        if inv and inv.file_path:
            from pathlib import Path
            if Path(inv.file_path).exists():
                source_file = {
                    "url": f"/invoices/{voucher.source_id}/file",
                    "name": Path(inv.file_path).name,
                }

    return JSONResponse({
        "success": True,
        "voucher": {
            "id": voucher.id,
            "voucher_no": voucher.voucher_no,
            "date": voucher.date.isoformat(),
            "voucher_word": voucher.voucher_word,
            "summary": voucher.summary or "",
            "status": voucher.status,
            "source_type": voucher.source_type,
            "creator": creator.display_name if creator else "",
            "reviewer": reviewer.display_name if reviewer else "",
        },
        "entries": [
            {
                "account_code": e.account_code or "",
                "account_name": e.account_name or "",
                "direction": e.direction,
                "amount": e.amount,
                "summary": e.summary or "",
            }
            for e in entries
        ],
        "source_file": source_file,
    })


@router.post("/renumber")
async def renumber_vouchers(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """重新编制当期凭证号：按凭证字分组，同字内按日期+ID排序"""
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return JSONResponse({"success": False, "msg": "无权限"})
    form = await request.form()
    period = form.get("period", "")
    if not period or "-" not in period:
        return JSONResponse({"success": False, "msg": "期间无效"})
    year_str, month_str = period.split("-")
    year, month = int(year_str), int(month_str)

    # 检查当期是否已结账
    from app.models.misc import ClosingPeriod
    cp = db.query(ClosingPeriod).filter(
        ClosingPeriod.company_id == user.company_id,
        ClosingPeriod.period == period,
    ).first()
    if cp and cp.is_closed:
        return JSONResponse({"success": False, "msg": f"期间 {period} 已结账，无法重新编号"})

    from app.services.voucher_service import batch_generate_numbers as bgn
    stats = bgn(db, user.company_id, year, month)
    if not stats:
        return JSONResponse({"success": False, "msg": "当期无凭证"})

    parts = [f"{w}字{count}张" for w, count in stats.items()]
    return JSONResponse({"success": True, "msg": f"重新编制完成：{'，'.join(parts)}"})


# ====== 凭证复制模块 ======

@router.get("/copy")
async def copy_voucher_page(request: Request, db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """复制凭证页面 - 选择来源期间浏览可复制的凭证"""
    if not user.company_id:
        return RedirectResponse(url="/dashboard", status_code=302)
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return RedirectResponse(url="/vouchers/", status_code=302)

    source_period = request.query_params.get("source_period", "")
    keyword = request.query_params.get("keyword", "")
    current_period = get_target_period(db, user.company_id)
    source_vouchers = []

    if source_period:
        query = db.query(Voucher).filter(
            Voucher.company_id == user.company_id,
            func.strftime("%Y-%m", Voucher.date) == source_period,
            Voucher.status.in_(["approved", "posted"]),
        )
        if keyword:
            query = query.filter(Voucher.summary.contains(keyword))
        source_vouchers = query.order_by(Voucher.date, Voucher.voucher_no).all()

    # 获取所有有数据的历史期间
    periods = db.query(
        func.strftime("%Y-%m", Voucher.date).label("period")
    ).filter(
        Voucher.company_id == user.company_id,
        Voucher.status.in_(["approved", "posted"]),
    ).distinct().order_by(func.strftime("%Y-%m", Voucher.date).desc()).all()
    period_list = [p[0] for p in periods]

    accounts = db.query(Account).filter(
        Account.company_id == user.company_id
    ).order_by(Account.code).all()

    return templates(request, "vouchers/copy.html", {
        "user": user, "periods": period_list,
        "source_period": source_period, "keyword": keyword,
        "source_vouchers": source_vouchers,
        "current_period": current_period,
        "accounts": accounts,
    })


@router.post("/copy/preview")
async def copy_preview(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """复制到预览 - 读取所选凭证完整信息，跳转到编辑确认界面"""
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return JSONResponse({"success": False, "msg": "无权限"})

    form = await request.form()
    ids_str_list = form.getlist("voucher_ids")
    if not ids_str_list:
        return JSONResponse({"success": False, "msg": "请选择要复制的凭证"})

    # 校验当期是否已结账
    from app.models.misc import ClosingPeriod
    current_period = get_target_period(db, user.company_id)
    cp = db.query(ClosingPeriod).filter(
        ClosingPeriod.company_id == user.company_id,
        ClosingPeriod.period == current_period,
    ).first()
    if cp and cp.is_closed:
        return JSONResponse({"success": False, "msg": "当前期间已结账，无法复制凭证"})

    # 读取所选凭证完整信息（存入session或直接渲染）
    ids = [int(x.strip()) for x in ids_str_list if x.strip()]
    source_vouchers = db.query(Voucher).filter(
        Voucher.id.in_(ids),
        Voucher.company_id == user.company_id,
        Voucher.status.in_(["approved", "posted"]),
    ).all()

    # 构建预览数据（JSON序列化）
    preview_list = []
    for sv in source_vouchers:
        entries = db.query(VoucherEntry).filter(
            VoucherEntry.voucher_id == sv.id
        ).order_by(VoucherEntry.sort_order).all()

        entry_list = []
        for e in entries:
            acct = db.query(Account).filter(Account.id == e.account_id).first()
            entry_list.append({
                "account_id": e.account_id,
                "account_code": e.account_code or (acct.code if acct else ""),
                "account_name": e.account_name or (acct.name if acct else ""),
                "direction": e.direction,
                "amount": e.amount,
                "summary": e.summary or "",
                "sort_order": e.sort_order,
            })

        preview_list.append({
            "source_id": sv.id,
            "source_no": sv.voucher_no,
            "date": current_period + "-01",  # 默认当期第一天
            "summary": sv.summary or "",
            "entries": entry_list,
        })

    import json
    accounts = db.query(Account).filter(Account.company_id == user.company_id).order_by(Account.code).all()
    return templates(request, "vouchers/copy_confirm.html", {
        "user": user,
        "preview_list": preview_list,
        "preview_json": json.dumps(preview_list, ensure_ascii=False),
        "accounts": accounts,
        "accounts_json": json.dumps([{"id": a.id, "code": a.code, "name": a.name} for a in accounts], ensure_ascii=False),
        "current_period": current_period,
    })


@router.post("/copy/save-draft")
async def copy_save_draft(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """保存复制预览为草稿（暂存功能）"""
    return await _batch_save_copied(request, db, user, finalize=False)


@router.post("/copy/confirm")
async def copy_confirm(request: Request, db: Session = Depends(get_db), user: User = Depends(require_active_company)):
    """确认生成所有复制凭证"""
    return await _batch_save_copied(request, db, user, finalize=True)


async def _batch_save_copied(request: Request, db: Session, user: User, finalize: bool) -> JSONResponse:
    """批量保存复制的凭证"""
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return JSONResponse({"success": False, "msg": "无权限"})

    form = await request.form()
    voucher_count = int(form.get("voucher_count", "0"))
    saved_count = 0
    errors = []

    for idx in range(voucher_count):
        idx_str = str(idx)
        summary = form.get(f"v_{idx}_summary", "")
        default_target = get_target_period(db, user.company_id) + "-01"
        date_str = form.get(f"v_{idx}_date", default_target)
        source_id = form.get(f"v_{idx}_source_id", "0")
        entry_count = int(form.get(f"v_{idx}_entries", "0"))

        try:
            voucher_date = date.fromisoformat(date_str)
        except ValueError:
            errors.append(f"凭证{idx+1}: 日期格式错误")
            continue

        # 生成新凭证号（使用源凭证的凭证字，或默认"记"）
        source_v = db.query(Voucher).filter(Voucher.id == int(source_id)).first() if source_id and source_id != "0" else None
        v_word = source_v.voucher_word if source_v else "记"
        voucher_no, seq = generate_voucher_no(db, user.company_id, v_word, voucher_date.year, voucher_date.month)

        # 解析分录
        entries_data = []
        for ei in range(entry_count):
            eid = f"v_{idx}_e_{ei}"
            account_id = int(form.get(f"{eid}_account_id", "0"))
            direction = form.get(f"{eid}_direction", "借")
            amount = float(form.get(f"{eid}_amount", "0"))
            entry_summary = form.get(f"{eid}_summary", "")
            if account_id and amount > 0:
                entries_data.append({
                    "account_id": account_id, "direction": direction,
                    "amount": amount, "summary": entry_summary,
                })

        if not entries_data:
            errors.append(f"凭证{idx+1}: 无有效分录")
            continue

        # 检查借贷平衡
        debit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "借")
        credit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "贷")
        if abs(debit_sum - credit_sum) > 0.01:
            if finalize:
                errors.append(f"凭证{idx+1}: 借贷不平衡（借={debit_sum:.2f}, 贷={credit_sum:.2f}）")
                continue

        voucher = Voucher(
            company_id=user.company_id,
            voucher_no=voucher_no,
            date=voucher_date,
            voucher_word=v_word,
            serial_no=seq,
            summary=summary,
            status="draft",
            source_type="copy",
            source_ref_id=int(source_id) if source_id and source_id != "0" else None,
            creator_id=user.id,
        )
        db.add(voucher)
        db.flush()

        for entry in entries_data:
            acct = db.query(Account).filter(Account.id == entry["account_id"]).first()
            db.add(VoucherEntry(
                voucher_id=voucher.id,
                account_id=entry["account_id"],
                account_code=acct.code if acct else "",
                account_name=acct.name if acct else "",
                direction=entry["direction"],
                amount=entry["amount"],
                summary=entry["summary"],
                sort_order=entry.get("sort_order", 0),
            ))

        saved_count += 1

    if errors:
        db.rollback()
        return JSONResponse({"success": False, "msg": " | ".join(errors)})

    db.commit()

    # 记录审计日志
    from app.models.misc import AuditLog
    db.add(AuditLog(
        company_id=user.company_id, user_id=user.id,
        username=user.username,
        action="voucher_copy",
        target_type="voucher",
        detail=f"{'确认生成' if finalize else '暂存'}复制凭证 {saved_count} 张",
    ))
    db.commit()

    msg = f"成功{'生成' if finalize else '暂存'}复制 {saved_count} 张凭证"
    return JSONResponse({"success": True, "msg": msg})


@router.post("/validate")
async def validate_voucher(
    request: Request,
    date_str: str = Form(...),
    summary: str = Form(""),
    voucher_word: str = Form("记"),
    db: Session = Depends(get_db),
    user: User = Depends(get_login_user),
):
    issues = []
    form = await request.form()
    try:
        date.fromisoformat(date_str)
    except ValueError:
        issues.append({"type": "date", "severity": "error", "message": "凭证日期格式无效", "suggestion": "请使用 YYYY-MM-DD 格式输入日期"})

    entries_data = []
    i = 0
    while f"account_id_{i}" in form:
        account_id = form.get(f"account_id_{i}")
        direction = form.get(f"direction_{i}")
        amount_str = form.get(f"amount_{i}", "0")
        try:
            amount = float(amount_str) if amount_str else 0
        except ValueError:
            amount = 0
        if account_id:
            entries_data.append({"account_id": int(account_id), "direction": direction, "amount": amount})
        i += 1

    if not entries_data:
        issues.append({"type": "no_entries", "severity": "error", "message": "凭证没有分录", "suggestion": "请至少添加一条分录"})
        return JSONResponse({"valid": False, "issues": issues})

    debit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "借")
    credit_sum = sum(e["amount"] for e in entries_data if e["direction"] == "贷")
    if abs(debit_sum - credit_sum) > 0.01:
        issues.append({"type": "balance", "severity": "error",
            "message": f"借贷不平衡：借方 {debit_sum:.2f} 贷方 {credit_sum:.2f}，差额 {debit_sum - credit_sum:.2f}",
            "suggestion": f"{'增加贷方' if debit_sum > credit_sum else '增加借方'}{abs(debit_sum - credit_sum):.2f}"})
    elif debit_sum == 0 and credit_sum == 0:
        issues.append({"type": "zero_amount", "severity": "warning", "message": "借贷金额均为 0", "suggestion": "请填写正确的金额"})
    else:
        issues.append({"type": "balance", "severity": "info", "message": f"借贷平衡：借方 {debit_sum:.2f} = 贷方 {credit_sum:.2f} ✅"})

    for idx, entry in enumerate(entries_data):
        n = idx + 1
        if entry["amount"] <= 0:
            issues.append({"type": "entry_amount", "severity": "error", "message": f"第 {n} 条分录金额为 {entry['amount']:.2f}", "suggestion": "金额必须大于 0"})
        if entry["account_id"]:
            acct = db.query(Account).filter(Account.id == entry["account_id"]).first()
            if not acct:
                issues.append({"type": "account_invalid", "severity": "error", "message": f"第 {n} 条分录的科目不存在", "suggestion": "请重新选择科目"})
            elif not acct.is_detail:
                issues.append({"type": "account_not_detail", "severity": "warning", "message": f"第 {n} 条分录「{acct.code} {acct.name}」不是末级科目", "suggestion": "请选择末级明细科目"})
        else:
            issues.append({"type": "account_missing", "severity": "error", "message": f"第 {n} 条分录未选择科目", "suggestion": "请选择科目"})

    if summary.strip():
        codes = set()
        for entry in entries_data:
            if entry["account_id"]:
                a = db.query(Account).filter(Account.id == entry["account_id"]).first()
                if a:
                    codes.add(a.code[:4])
        from app.services.scene_service import get_scene_rules
        for rule in get_scene_rules(user.company_id or 0, db):
            if not rule.keywords:
                continue
            kw_list = [k.strip() for k in rule.keywords.split(",")]
            if any(kw in summary for kw in kw_list if kw):
                if not any(c.startswith(rule.debit_account_code) for c in codes):
                    issues.append({"type": "account_mismatch", "severity": "warning", "message": f"摘要「{summary}」建议借方 {rule.debit_account_code}XX，当前未匹配", "suggestion": "请检查借方科目"})
                if not any(c.startswith(rule.credit_account_code) for c in codes):
                    issues.append({"type": "account_mismatch", "severity": "warning", "message": f"摘要「{summary}」建议贷方 {rule.credit_account_code}XX，当前未匹配", "suggestion": "请检查贷方科目"})
                break

    has_err = any(i["severity"] == "error" for i in issues)
    return JSONResponse({"valid": not has_err, "issues": issues})


@router.get("/export")
async def export_vouchers(
    request: Request,
    period: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_login_user),
):
    """导出指定期间的所有凭证到Excel"""
    if not user.company_id:
        return Response("无权限")
    from app.services.period_helper import get_current_period
    period = period or get_current_period(request, db, user.company_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"凭证导出_{period}"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 标题行
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"凭证导出 - {period}"
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    # 表头
    headers = ["凭证号", "日期", "凭证字", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    # 查询已过账凭证
    vouchers = db.query(Voucher).filter(
        Voucher.company_id == user.company_id,
        func.strftime("%Y-%m", Voucher.date) == period,
        Voucher.status == "posted",
    ).order_by(Voucher.voucher_no).all()

    row = 4
    for v in vouchers:
        entries = db.query(VoucherEntry).filter(
            VoucherEntry.voucher_id == v.id
        ).order_by(VoucherEntry.sort_order).all()
        for e in entries:
            ws.cell(row=row, column=1, value=v.voucher_no).border = thin
            ws.cell(row=row, column=2, value=v.date.isoformat()).border = thin
            ws.cell(row=row, column=3, value=v.voucher_word or "").border = thin
            ws.cell(row=row, column=4, value=e.summary or v.summary or "").border = thin
            ws.cell(row=row, column=5, value=e.account_code or "").border = thin
            ws.cell(row=row, column=6, value=e.account_name or "").border = thin
            debit = e.amount if e.direction == "借" else ""
            credit = e.amount if e.direction == "贷" else ""
            ws.cell(row=row, column=7, value=debit).border = thin
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=credit).border = thin
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            row += 1

    # 列宽
    widths = [20, 12, 8, 30, 12, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    filename = f"vouchers_{period}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )


@router.get("/import-template")
async def import_template(db: Session = Depends(get_db), user: User = Depends(get_login_user)):
    """下载凭证导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "凭证导入模板"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    headers = ["日期", "凭证字", "摘要", "科目编码", "借方金额", "贷方金额"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    # 示例数据
    samples = [
        ["2026-06-05", "转", "计提6月工资", "5602", 40000, ""],
        ["2026-06-05", "转", "计提6月工资", "2211", "", 40000],
        ["2026-06-10", "付", "支付办公用品", "5602", 1500, ""],
        ["2026-06-10", "付", "支付办公用品", "1002", "", 1500],
        ["2026-06-22", "收", "销售收入", "1002", 80000, ""],
        ["2026-06-22", "收", "销售收入", "5001", "", 80000],
    ]
    for i, row_data in enumerate(samples, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin
            if col >= 5:
                cell.number_format = '#,##0.00'

    widths = [14, 8, 28, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=voucher_import_template.xlsx"},
    )


@router.post("/import")
async def import_vouchers(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_active_company),
):
    """批量导入凭证（Excel格式）"""
    if not user.company_id:
        return JSONResponse({"success": False, "msg": "无权限"})
    if user.role not in ("inputer", "company_admin", "super_admin"):
        return JSONResponse({"success": False, "msg": "无导入权限"})

    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        return JSONResponse({"success": False, "msg": "请选择文件"})

    import io, openpyxl
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        return JSONResponse({"success": False, "msg": f"文件解析失败: {str(e)}"})

    from datetime import date as dt_date
    from app.services.voucher_service import generate_voucher_no, post_voucher
    from app.models import Account

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success_count = 0
    fail_count = 0
    errors = []

    # 按 (日期, 凭证字, 摘要) 分组构建凭证
    voucher_groups = {}
    for row in rows:
        if not row or not row[0]:
            continue
        raw_date = str(row[0]).strip()[:10] if row[0] else ""
        word = str(row[1]).strip() if row[1] and str(row[1]).strip() else "转"
        summary = str(row[2]).strip() if row[2] else ""
        acct_code = str(row[3]).strip() if row[3] else ""
        debit = float(row[4]) if row[4] and str(row[4]).strip() else 0.0
        credit = float(row[5]) if row[5] and str(row[5]).strip() else 0.0

        if not raw_date or not acct_code:
            continue
        try:
            vd = dt_date.fromisoformat(raw_date)
        except ValueError:
            errors.append(f"日期格式错误: {raw_date}")
            fail_count += 1
            continue

        key = (raw_date, word, summary)
        if key not in voucher_groups:
            voucher_groups[key] = {"date": vd, "word": word, "summary": summary, "entries": []}
        direction = "借" if debit and debit > 0 else "贷"
        amount = debit or credit
        voucher_groups[key]["entries"].append((acct_code, direction, amount, summary))

    # 逐个创建凭证
    for key, vg in voucher_groups.items():
        entries = vg["entries"]
        # 校验借贷平衡
        total_debit = sum(e[2] for e in entries if e[1] == "借")
        total_credit = sum(e[2] for e in entries if e[1] == "贷")
        if abs(total_debit - total_credit) > 0.01:
            errors.append(f"借贷不平衡: {key[2]} (借={total_debit:.2f} 贷={total_credit:.2f})")
            fail_count += 1
            continue

        try:
            vd = vg["date"]
            voucher_no, seq = generate_voucher_no(db, user.company_id, vg["word"], vd.year, vd.month)
            creator_id = user.id
            voucher = Voucher(
                company_id=user.company_id, voucher_no=voucher_no, date=vd,
                voucher_word=vg["word"], serial_no=seq, summary=vg["summary"],
                status="draft", source_type="manual", creator_id=creator_id,
            )
            db.add(voucher)
            db.flush()
            for idx, (code, direction, amount, esummary) in enumerate(entries):
                acct = db.query(Account).filter(
                    Account.company_id == user.company_id, Account.code == code
                ).first()
                if not acct:
                    acct = db.query(Account).filter(
                        Account.company_id == user.company_id, Account.code.like(f"{code}%")
                    ).first()
                aid = acct.id if acct else 0
                aname = acct.name if acct else code
                db.add(VoucherEntry(
                    voucher_id=voucher.id, account_id=aid,
                    account_code=code, account_name=aname,
                    direction=direction, amount=amount,
                    summary=esummary or vg["summary"], sort_order=idx,
                ))
            db.commit()

            # 提交→审核→过账（完整流程）
            reviewer_user = db.query(User).filter(
                User.company_id == user.company_id, User.role == "reviewer"
            ).first()
            reviewer_id = reviewer_user.id if reviewer_user else creator_id
            voucher.status = "pending"
            db.commit()
            voucher.status = "approved"
            voucher.reviewer_id = reviewer_id
            voucher.reviewed_at = datetime.now()
            db.commit()
            ok, msg = post_voucher(voucher.id, reviewer_id, db)
            if ok:
                success_count += 1
            else:
                errors.append(f"过账失败 {voucher_no}: {msg}")
                fail_count += 1
        except Exception as e:
            db.rollback()
            errors.append(f"导入失败 {vg['summary']}: {str(e)}")
            fail_count += 1

    return JSONResponse({
        "success": True,
        "msg": f"导入完成：成功 {success_count} 张，失败 {fail_count} 张",
        "errors": errors[:10],
    })
