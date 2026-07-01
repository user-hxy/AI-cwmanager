"""报表生成服务"""
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models import Account, AccountBalance, Voucher, VoucherEntry, Company
from datetime import date
from calendar import monthrange
from collections import defaultdict


def _get_opening_data(company_id: int, period: str, db: Session) -> dict:
    """获取期初数据：优先取本期的opening_balance，没有则取上期closing_balance"""
    balances = db.query(AccountBalance).filter(
        AccountBalance.company_id == company_id,
        AccountBalance.period == period,
    ).all()
    if balances:
        return {b.account_id: b.opening_balance for b in balances}

    # 没有本期记录，找最近的上一期
    year, month = int(period[:4]), int(period[5:7])
    for _ in range(60):  # 最多往前找5年
        month -= 1
        if month == 0:
            month = 12
            year -= 1
        prev = f"{year}-{month:02d}"
        prev_bals = db.query(AccountBalance).filter(
            AccountBalance.company_id == company_id,
            AccountBalance.period == prev,
        ).all()
        if prev_bals:
            return {b.account_id: b.closing_balance for b in prev_bals}
    return {}


def _get_period_voucher_activity(company_id: int, start_period: str, end_period: str, db: Session) -> tuple:
    """从凭证分录计算期间借贷发生额（不依赖AccountBalance表）
    返回 (debit_by_account, credit_by_account)
    """
    start_date = date(int(start_period[:4]), int(start_period[5:7]), 1)
    ey, em = int(end_period[:4]), int(end_period[5:7])
    end_date = date(ey, em, monthrange(ey, em)[1])

    entries = db.query(VoucherEntry, Voucher).join(
        Voucher, VoucherEntry.voucher_id == Voucher.id
    ).filter(
        Voucher.company_id == company_id,
        Voucher.date >= start_date,
        Voucher.date <= end_date,
        Voucher.status != "draft",
    ).all()

    debit = defaultdict(float)
    credit = defaultdict(float)
    for entry, _ in entries:
        if entry.direction == "借":
            debit[entry.account_id] += entry.amount
        else:
            credit[entry.account_id] += entry.amount
    return debit, credit


def get_trial_balance(company_id: int, start_period: str, end_period: str, db: Session) -> list:
    """科目余额汇总表 — 从凭证分录计算本期发生额，AccountBalance仅用于期初期末"""
    # 期初数据
    start_data = _get_opening_data(company_id, start_period, db)

    # 期间借贷发生额（从凭证分录直接计算）
    period_debit, period_credit = _get_period_voucher_activity(company_id, start_period, end_period, db)

    # 期末余额：优先取AccountBalance，否则从期初+本期发生计算
    end_balances = db.query(AccountBalance).filter(
        AccountBalance.company_id == company_id,
        AccountBalance.period == end_period,
    ).all()
    end_map = {b.account_id: b for b in end_balances}

    # 获取所有科目
    accounts = db.query(Account).filter(
        Account.company_id == company_id,
    ).order_by(Account.code).all()

    result = []
    for acct in accounts:
        opening = start_data.get(acct.id, 0)
        debit = period_debit.get(acct.id, 0)
        credit = period_credit.get(acct.id, 0)

        end = end_map.get(acct.id)
        if end:
            closing = end.closing_balance
        else:
            # 从期初+本期发生计算期末
            if acct.direction == "借":
                closing = opening + debit - credit
            else:
                closing = opening + credit - debit

        # 贷方方向科目取反展示（避免 -0.00 显示问题）
        if acct.direction == "贷":
            opening = -opening if opening else 0.0
            closing = -closing if closing else 0.0

        # 只显示有变化的科目
        if opening != 0 or debit != 0 or credit != 0 or closing != 0:
            result.append({
                "account_code": acct.code, "account_name": acct.name,
                "category": acct.category,
                "opening_balance": opening,
                "debit_amount": debit, "credit_amount": credit,
                "closing_balance": closing,
            })
    return sorted(result, key=lambda x: x["account_code"])


def generate_balance_sheet(company_id: int, start_period: str, end_period: str, db: Session) -> dict:
    """资产负债表 — 标准格式"""
    sheet = {
        "title": "资产负债表", "company": "",
        "period": f"{end_period}",
        "assets": [], "liabilities": [], "equity": [],
        "total_assets": 0, "total_liabilities": 0, "total_equity": 0,
    }
    comp = db.query(Company).filter(Company.id == company_id).first()
    if comp:
        sheet["company"] = comp.name

    # 期初数据
    start_data = _get_opening_data(company_id, start_period, db)

    # 期间借贷发生额
    period_debit, period_credit = _get_period_voucher_activity(company_id, start_period, end_period, db)

    # 期末余额：优先取AccountBalance
    end_balances = db.query(AccountBalance).filter(
        AccountBalance.company_id == company_id,
        AccountBalance.period == end_period,
    ).all()
    end_map = {b.account_id: b for b in end_balances}

    accounts = db.query(Account).filter(
        Account.company_id == company_id,
    ).order_by(Account.code).all()

    for acct in accounts:
        bal = end_map.get(acct.id)
        if bal:
            closing = bal.closing_balance
        else:
            opening = start_data.get(acct.id, 0)
            debit = period_debit.get(acct.id, 0)
            credit = period_credit.get(acct.id, 0)
            if acct.direction == "借":
                closing = opening + debit - credit
            else:
                closing = opening + credit - debit

        opening = start_data.get(acct.id, 0)

        # 按科目方向调整符号（避免 -0.00 显示问题）
        if acct.direction == "贷":
            closing = -closing if closing else 0.0
            opening = -opening if opening else 0.0

        item = {"code": acct.code, "name": acct.name, "opening": opening, "closing": closing}
        if acct.category == "资产":
            sheet["assets"].append(item)
            sheet["total_assets"] += closing
            sheet["total_assets_opening"] = sheet.get("total_assets_opening", 0) + opening
        elif acct.category == "负债":
            sheet["liabilities"].append(item)
            sheet["total_liabilities"] += closing
            sheet["total_liabilities_opening"] = sheet.get("total_liabilities_opening", 0) + opening
        elif acct.category == "权益":
            sheet["equity"].append(item)
            sheet["total_equity"] += closing
            sheet["total_equity_opening"] = sheet.get("total_equity_opening", 0) + opening

    sheet.setdefault("total_assets_opening", 0)
    sheet.setdefault("total_liabilities_opening", 0)
    sheet.setdefault("total_equity_opening", 0)
    return sheet


def generate_income_statement(company_id: int, start_period: str, end_period: str, db: Session) -> dict:
    """利润表 — 从VoucherEntry直接汇总（排除结转凭证），按科目方向取净额"""
    statement = {
        "title": "利润表", "company": "",
        "period": f"{start_period} 至 {end_period}",
        "revenue_items": [], "expense_items": [],
        "total_revenue": 0, "total_expense": 0, "net_profit": 0,
    }
    comp = db.query(Company).filter(Company.id == company_id).first()
    if comp:
        statement["company"] = comp.name

    accounts = db.query(Account).filter(
        Account.company_id == company_id, Account.category == "损益",
    ).order_by(Account.code).all()

    # 直接用日期范围过滤（避免func.strftime兼容性问题）
    start_date = date(int(start_period[:4]), int(start_period[5:7]), 1)
    ey, em = int(end_period[:4]), int(end_period[5:7])
    end_date = date(ey, em, monthrange(ey, em)[1])

    entries = db.query(VoucherEntry, Voucher).join(
        Voucher, VoucherEntry.voucher_id == Voucher.id
    ).filter(
        Voucher.company_id == company_id,
        Voucher.date >= start_date,
        Voucher.date <= end_date,
        Voucher.source_type != "carry_forward",
        Voucher.status != "draft",
    ).all()

    # 按科目和方向汇总金额
    amt_by_acct = defaultdict(lambda: {"借": 0.0, "贷": 0.0})
    for entry, _ in entries:
        amt_by_acct[entry.account_id][entry.direction] += entry.amount

    for acct in accounts:
        amts = amt_by_acct.get(acct.id, {"借": 0.0, "贷": 0.0})
        debit = amts["借"]
        credit = amts["贷"]

        if acct.direction == "贷":
            net = credit - debit
        else:
            net = debit - credit

        amount = abs(net) if net > 0 else 0
        direction = "收入" if acct.code in ("5001", "5051", "5111", "5301") else "费用"

        item = {"code": acct.code, "name": acct.name, "amount": amount, "direction": direction}
        if direction == "收入":
            statement["revenue_items"].append(item)
            statement["total_revenue"] += amount
        else:
            statement["expense_items"].append(item)
            statement["total_expense"] += amount

    statement["net_profit"] = statement["total_revenue"] - statement["total_expense"]
    return statement


def generate_cash_flow(company_id: int, start_period: str, end_period: str, db: Session) -> dict:
    """现金流量表（简化版）"""
    return {
        "title": "现金流量表", "company": "",
        "period": f"{start_period} 至 {end_period}",
        "message": "现金流量表需要基于现金流水详细数据生成，当前为简化版本",
        "items": [],
    }


def export_to_excel_standard(data: dict, report_type: str) -> bytes:
    """将标准格式报表导出为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("title", "报表")
    ws.merge_cells("A1:E1")
    c = ws["A1"]; c.value = data.get("title", ""); c.font = Font(bold=True, size=14); c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")
    c = ws["A2"]; c.value = f"{data.get('company', '')} - {data.get('period', '')}"; c.alignment = Alignment(horizontal="center")
    headers = ["项目", "行次", "期末余额", "期初余额"]
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col); c.value = h; c.font = Font(bold=True); c.alignment = Alignment(horizontal="center"); c.border = thin
    row = 5
    for section_key in ["left_items"]:
        for item in data.get(section_key, []):
            is_h = item.get("is_header", False)
            ws.cell(row=row, column=1, value=item.get("name", ""))
            ws.cell(row=row, column=2, value=item.get("line_no", ""))
            closing = item.get("closing", "")
            opening = item.get("opening", "")
            ws.cell(row=row, column=3, value=f"{closing:.2f}" if isinstance(closing, (int, float)) else "")
            ws.cell(row=row, column=4, value=f"{opening:.2f}" if isinstance(opening, (int, float)) else "")
            if is_h:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_to_excel(data: dict, report_type: str) -> bytes:
    """将报表数据导出为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("title", "报表")
    ws.merge_cells("A1:D1")
    c = ws["A1"]; c.value = data.get("title", ""); c.font = Font(bold=True, size=14); c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    c = ws["A2"]; c.value = f"{data.get('company', '')} - {data.get('period', '')}"; c.alignment = Alignment(horizontal="center")
    headers = ["科目编码", "科目名称", "期初余额", "期末余额"]
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col); c.value = h; c.font = Font(bold=True); c.alignment = Alignment(horizontal="center"); c.border = thin
    row = 5
    for section_key in ["assets", "liabilities", "equity", "revenue_items", "expense_items"]:
        for item in data.get(section_key, []):
            ws.cell(row=row, column=1, value=item.get("code", ""))
            ws.cell(row=row, column=2, value=item.get("name", ""))
            ws.cell(row=row, column=3, value=item.get("opening", item.get("amount", 0)))
            ws.cell(row=row, column=4, value=item.get("closing", 0))
            row += 1
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
